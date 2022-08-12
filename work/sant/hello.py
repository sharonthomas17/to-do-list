import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook

from_add="1998wlinterfell@gmail.com"
password="uqptxtkqjamxnwbj"

msg=MIMEMultipart()
msg['from']=from_add
wb= load_workbook(filename='details.xlsx')
sheet=wb.active
nrows=sheet.max_row
ncols=sheet.max_column
he=[]
for i in range(1,ncols+1):
    cell=sheet.cell(row=1,column=i)
    h=cell.value
    he.append(h)
na=he.index('NAME')
m=he.index('MARK')
mai=he.index('EMAIL_ID')
sta=he.index('STATUS')
    
for i in range(2,nrows+1):
    cell=sheet.cell(row=i,column=na+1)
    n=cell.value   
    subject=f'EXAM RESULT OF {n}'
    
    cell=sheet.cell(row=i,column=m+1)
    m1=cell.value

    cell=sheet.cell(row=i,column=sta+1)
    if m1>35:
        st="PASSED"
    else:
        st="FAILED"    
    cell.value=st
    body=f'Hi {n},You Have {st} our exam.\n\n your mark is {m1} '
    cell=sheet.cell(row=i,column=mai+1)
    to_add= cell.value    
    wb.save('details.xlsx')
    msg=MIMEMultipart()
    msg['from']=from_add
    msg['subject']=subject
    msg['to']=to_add
    print('mail sent to:',to_add)

    msg.attach(MIMEText(body,'plain'))
   
    filename='Result.png'
    attachment=open(filename,'rb')

    p= MIMEBase('image','png')  

    p.set_payload(attachment.read())
 
    encoders.encode_base64(p)

    p.add_header('content-Disposition',"attachment;filename=%s"%filename)

    msg.attach(p)

    server=smtplib.SMTP('smtp.gmail.com',587)
    server.starttls()
    server.login(from_add,password)
    server.send_message(msg)
