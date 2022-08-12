import smtplib
from openpyxl import load_workbook

with smtplib.SMTP('smtp.gmail.com',587) as smtp:
    smtp.starttls()

    smtp.login("1998wlinterfell@gmail.com","uqptxtkqjamxnwbj")
    wb= load_workbook(filename='details.xlsx')
    sheet=wb.active
    nrows=sheet.max_row
    ncols=sheet.max_column
    he=[]
    for i in range(1,ncols+1):
        cell=sheet.cell(row=1,column=i)
        h=cell.value
        he.append(h)
    na=he.index('NAME')
    m=he.index('MARK')
    mai=he.index('EMAIL_ID')
    sta=he.index('STATUS')
    
    for i in range(2,nrows+1):
        cell=sheet.cell(row=i,column=na+1)
        n=cell.value   
        subject=f'EXAM RESULT'
    
        cell=sheet.cell(row=i,column=m+1)
        m1=cell.value

        cell=sheet.cell(row=i,column=sta+1)
        if m1>35:
            st="PASSED"
        else:
            st="FAILED"    
        cell.value=st
        body=f'Hi {n},You Have {st} our exam.\n\n your mark is {m1} '
        

        msg=f'subject:{subject}\n\n{body}'

        cell=sheet.cell(row=i,column=mai+1)
        li= cell.value
        print('Mail sent to:',li)
        wb.save('details.xlsx')
        smtp.sendmail('1998wlinterfell@gmail.com',li,msg)
    